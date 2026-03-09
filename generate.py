"""
EarlySign · generate.py
-----------------------
Lit le fichier Excel depuis ./input/
Régénère index.html
Fait git add + commit + push automatiquement

Usage :
    python generate.py
"""

import os, sys, glob, json, subprocess
from pathlib import Path
from datetime import datetime
import openpyxl

# ── CONFIG ────────────────────────────────────────────────────
INPUT_DIR   = Path(__file__).parent / "input"
OUTPUT_HTML = Path(__file__).parent / "index.html"
MOIS = ["janvier","février","mars","avril","mai","juin",
        "juillet","août","septembre","octobre","novembre","décembre"]

def to_fr(dt):
    if isinstance(dt, datetime):
        return f"{dt.day} {MOIS[dt.month-1]} {dt.year}"
    return str(dt)

def today_fr():
    d = datetime.today()
    return f"{d.day} {MOIS[d.month-1]} {d.year}"

# ── LECTURE EXCEL ─────────────────────────────────────────────
def find_excel():
    files = list(INPUT_DIR.glob("*.xlsx")) + list(INPUT_DIR.glob("*.xlsm"))
    if not files:
        print("❌  Aucun fichier .xlsx trouvé dans ./input/")
        sys.exit(1)
    # Prend le plus récent
    return sorted(files, key=lambda f: f.stat().st_mtime)[-1]

def parse_excel(path):
    print(f"📂  Lecture : {path.name}")
    wb = openpyxl.load_workbook(path, data_only=True)
    data = {}

    # ── Onglet 1 : Référentiel ────────────────────────────────
    ws1 = wb["1_Referentiel"]
    rows1 = [r for r in ws1.iter_rows(values_only=True) if any(c is not None for c in r)]
    # Titre chantier
    titre = str(rows1[0][0]) if rows1 else ""
    chantier = ""
    if "Chantier :" in titre:
        chantier = titre.split("Chantier :")[-1].strip()
        if chantier.endswith(")"):
            chantier = chantier.rsplit("(", 1)[0].strip()
    data["chantier"] = chantier or "Chantier"

    # ── Onglet 2 : Saisie Terrain ─────────────────────────────
    ws2 = wb["2_Saisie_Terrain"]
    rows2 = [r for r in ws2.iter_rows(values_only=True) if any(c is not None for c in r)]

    # Semaine & date
    semaine, date_saisie = "", ""
    for r in rows2:
        if r[0] and "Semaine" in str(r[0]):
            semaine = str(r[1] or "")
            raw_date = r[3]
            if isinstance(raw_date, datetime):
                date_saisie = to_fr(raw_date)
            elif raw_date:
                date_saisie = str(raw_date)
            break
    data["semaine"]     = semaine
    data["date_saisie"] = date_saisie or today_fr()

    # Lignes d'équipements (après la ligne d'en-tête)
    header_found = False
    equip_rows = []
    h_totales_terrain = 0.0
    h_pointees = 0.0

    for r in rows2:
        if r[0] and str(r[0]).startswith("Zone") and r[1] == "Lot":
            header_found = True
            continue
        if not header_found:
            continue
        # Saut de ligne zone ou ligne de légende
        if r[0] and (str(r[0]).startswith("  ZONE") or str(r[0]).startswith("CFO ")):
            continue
        # Total heures
        if r[0] and "TOTAL" in str(r[0]):
            h_totales_terrain = float(r[9] or 0)
            continue
        if r[0] and "pointées" in str(r[0]).lower():
            h_pointees = float(r[9] or 0)
            continue
        if r[0] and "Écart" in str(r[0]):
            continue
        # Ligne équipement
        if r[0] and r[1] in ("CFO", "CFA"):
            equip_rows.append({
                "zone": str(r[0]),
                "lot":  str(r[1]),
                "ref":  str(r[2] or ""),
                "desig": str(r[3] or ""),
                "qte":  float(r[4] or 0),
                "s1":   float(r[5] or 0),
                "s2":   float(r[6] or 0),
                "s3":   float(r[7] or 0),
                "pct":  float(r[8] or 0),
                "h_theo": float(r[9] or 0),
            })

    data["equip_rows"]        = equip_rows
    data["h_totales_terrain"] = h_totales_terrain
    data["h_pointees"]        = h_pointees

    # ── Onglet 3 : Contrôle Cohérence ────────────────────────
    ws3 = wb["3_Controle_Coherence"]
    rows3 = [r for r in ws3.iter_rows(values_only=True) if any(c is not None for c in r)]

    coher_heures = []
    coher_bl     = []
    signal_global = ""
    synthese = {}

    section = None
    for r in rows3:
        if not r[0]: continue
        v = str(r[0])
        if "CONTRÔLE 1" in v: section = "h"; continue
        if "CONTRÔLE 2" in v: section = "bl"; continue
        if "SYNTHÈSE"   in v: section = "synth"; continue
        if v in ("Zone", "Zone / Équipement", "Indicateur"): continue

        if section == "h" and r[1] is not None:
            try:
                coher_heures.append({
                    "zone":      v,
                    "theo":      float(r[1] or 0),
                    "point":     float(r[2] or 0),
                    "ecart_h":   float(r[3] or 0),
                    "ecart_pct": float(r[4] or 0),
                    "statut":    str(r[5] or ""),
                    "action":    str(r[7] or ""),
                })
            except: pass

        elif section == "bl" and r[1] is not None:
            try:
                coher_bl.append({
                    "item":   v,
                    "decl":   float(r[1] or 0),
                    "livre":  float(r[2] or 0),
                    "ecart":  float(r[3] or 0),
                    "statut": str(r[4] or ""),
                    "action": str(r[5] or ""),
                })
            except: pass

        elif section == "synth":
            if r[1]:
                synthese[v] = {"valeur": str(r[1]), "detail": str(r[2] or "")}

    data["coher_heures"]  = coher_heures
    data["coher_bl"]      = coher_bl
    data["synthese"]      = synthese

    # ── Onglet 4 : Dashboard DR ───────────────────────────────
    ws4 = wb["4_Dashboard_DR"]
    rows4 = [r for r in ws4.iter_rows(values_only=True) if any(c is not None for c in r)]

    # Titre (semaine)
    titre4 = str(rows4[0][0]) if rows4 else ""
    sem4 = ""
    if "Semaine" in titre4:
        sem4 = titre4.split("Semaine")[-1].strip()
    if sem4 and not semaine:
        data["semaine"] = sem4

    # KPIs ligne 4 (index 3)
    kpi_vals = {}
    kpi_labels = {}
    for i, r in enumerate(rows4):
        if r[0] and "Avancement" in str(r[0]) and r[2] and "Heures" in str(r[2]):
            # En-tête KPI
            kpi_labels = {
                "avanc": str(r[0]).replace("\n"," "),
                "h_theo": str(r[2]).replace("\n"," "),
                "h_point": str(r[4]).replace("\n"," "),
                "derives": str(r[6]).replace("\n"," "),
                "bl": str(r[8]).replace("\n"," "),
            }
            # Valeurs ligne suivante
            if i+1 < len(rows4):
                nxt = rows4[i+1]
                kpi_vals = {
                    "avanc":   str(nxt[0] or ""),
                    "h_theo":  str(nxt[2] or ""),
                    "h_point": str(nxt[4] or ""),
                    "derives": str(nxt[6] or ""),
                    "bl":      str(nxt[8] or ""),
                }
            break

    data["kpi"] = kpi_vals

    # Tableau zones
    zones = []
    zone_header = False
    for r in rows4:
        if r[0] and "Zone" == str(r[0]).strip() and r[1] and "CFO" in str(r[1]):
            zone_header = True
            continue
        if not zone_header: continue
        if r[0] and "RECOMMANDATIONS" in str(r[0]): break
        if r[0] and r[1] is not None:
            zones.append({
                "zone":    str(r[0]),
                "cfo":     str(r[1] or ""),
                "cfa":     str(r[2] or ""),
                "moy":     str(r[3] or ""),
                "h_cfo":   str(r[4] or ""),
                "h_cfa":   str(r[5] or ""),
                "h_point": str(r[6] or ""),
                "ecart":   str(r[7] or ""),
                "bl":      str(r[8] or ""),
                "signal":  str(r[9] or ""),
            })

    data["zones"] = zones

    # Recommandations
    recomms = []
    in_recomm = False
    for r in rows4:
        if r[0] and "RECOMMANDATIONS" in str(r[0]):
            in_recomm = True
            continue
        if not in_recomm: continue
        if r[0] and str(r[0]) == "Priorité": continue
        if r[0] and r[1] and r[2]:
            recomms.append({
                "prio":  str(r[0]),
                "zone":  str(r[1]),
                "action": str(r[2]),
            })
    data["recommandations"] = recomms

    # Avancement numérique
    avanc_str = kpi_vals.get("avanc", "0%").replace("%","").strip()
    try:
        data["avancement"] = float(avanc_str)
    except:
        data["avancement"] = 0.0

    return data

# ── HELPERS HTML ──────────────────────────────────────────────
def signal_class(s):
    s = s.upper()
    if "ALERTE" in s or "URGENT" in s: return "err"
    if "VIGILANCE" in s or "ÉCART" in s or "SUIVRE" in s: return "warn"
    return "ok"

def signal_label(s):
    s_up = s.upper()
    if "ALERTE" in s_up:    return "🔴 ALERTE"
    if "VIGILANCE" in s_up: return "🟠 VIGILANCE"
    if "ÉCART" in s_up:     return "🟠 ÉCART"
    return "✅ OK"

def prio_class(p):
    p = p.upper()
    if "URGENT" in p or "🔴" in p: return "err"
    if "SUIVRE" in p or "🟠" in p: return "warn"
    return "ok"

def pct_num(s):
    """'52%' → 52"""
    try: return float(str(s).replace("%","").replace("+","").replace(" ",""))
    except: return 0

# ── BUILD HTML ────────────────────────────────────────────────
def build_html(data):
    av = data["avancement"]
    kpi = data.get("kpi", {})
    zones = data.get("zones", [])
    recomms = data.get("recommandations", [])
    coher_h = data.get("coher_heures", [])
    coher_bl = data.get("coher_bl", [])
    synth = data.get("synthese", {})
    sem = data.get("semaine", "")
    sem_label = f"Semaine {sem}" if sem else ""

    # ── Zones table rows ──────────────────────────────────────
    zone_rows_html = ""
    for z in zones:
        cfo_n = pct_num(z["cfo"])
        cfa_n = pct_num(z["cfa"])
        moy_n = pct_num(z["moy"])
        ecart_n = pct_num(z["ecart"])
        ecart_color = "var(--err)" if abs(ecart_n) > 15 else ("var(--warn)" if abs(ecart_n) > 5 else "var(--ok)")
        sc = signal_class(z["signal"])
        sl = signal_label(z["signal"])
        zone_rows_html += f"""
        <tr>
          <td>{z['zone']}</td>
          <td class="pct-cell">{z['cfo']}
            <span class="mini-bar-wrap"><span class="mini-bar" style="width:{min(cfo_n,100)}%;background:var(--p)"></span></span>
          </td>
          <td class="pct-cell">{z['cfa']}
            <span class="mini-bar-wrap"><span class="mini-bar" style="width:{min(cfa_n,100)}%;background:var(--p-m)"></span></span>
          </td>
          <td class="pct-cell" style="color:var(--p)">{z['moy']}</td>
          <td>{z['h_cfo']}</td>
          <td>{z['h_cfa']}</td>
          <td>{z['h_point']}</td>
          <td style="font-weight:700;color:{ecart_color}">{z['ecart']}</td>
          <td>{z['bl']}</td>
          <td><span class="sig {sc}">{sl}</span></td>
        </tr>"""

    # ── Alertes ───────────────────────────────────────────────
    alertes_html = ""
    for r in recomms:
        pc = prio_class(r["prio"])
        icon = "🔴" if pc == "err" else ("🟠" if pc == "warn" else "✅")
        title = r["prio"].replace("🔴","").replace("🟠","").replace("✅","").strip()
        alertes_html += f"""
        <div class="alerte {pc}">
          <div class="alerte-icon">{icon}</div>
          <div class="alerte-body">
            <div class="alerte-title">{title} — {r['zone']}</div>
            <div class="alerte-desc">{r['action']}</div>
          </div>
        </div>"""

    # ── Cohérence heures ──────────────────────────────────────
    coher_h_html = ""
    for r in coher_h:
        ep = r["ecart_pct"]
        sign = "+" if ep > 0 else ""
        ec = "var(--err)" if abs(ep) > 15 else "var(--ink-70)"
        icon = "✅" if "OK" in r["statut"] else "🔴"
        coher_h_html += f"""
        <div class="coher-row">
          <span class="coher-label">{r['zone']}</span>
          <span class="coher-val" style="color:{ec}">{icon} {sign}{ep}%</span>
        </div>"""

    # ── Cohérence BL ──────────────────────────────────────────
    coher_bl_html = ""
    for r in coher_bl:
        is_ok = "OK" in r["statut"]
        icon  = "✅" if is_ok else "🟠"
        color = "var(--ink-70)" if is_ok else "var(--warn)"
        coher_bl_html += f"""
        <div class="coher-row">
          <span class="coher-label">{r['item']}</span>
          <span class="coher-val" style="color:{color}">{icon} {int(r['decl'])} / {int(r['livre'])}</span>
        </div>"""

    # ── Synthèse KPIs ─────────────────────────────────────────
    def synth_kpi(key, lbl, fallback_val, fallback_sub):
        if key in synth:
            val = synth[key]["valeur"]
            sub = synth[key]["detail"]
            cl  = signal_class(val + sub)
        else:
            val, sub, cl = fallback_val, fallback_sub, signal_class(fallback_val)
        css = "err" if cl == "err" else ("warn" if cl == "warn" else "ok")
        return f"""
        <div class="kpi">
          <div class="lbl">{lbl}</div>
          <div class="val {css}">{val}</div>
          <div class="sub">{sub}</div>
        </div>"""

    synth_html  = synth_kpi("Zones avec dérive heures détectée","Zones avec dérive heures", kpi.get("derives","—"),"")
    synth_html += synth_kpi("Zones avec écart BL matériel","Écarts BL matériel", kpi.get("bl","—"),"")
    if "Zones cohérentes — RAS" in synth:
        s = synth["Zones cohérentes — RAS"]
        synth_html += f'<div class="kpi"><div class="lbl">Zones cohérentes</div><div class="val ok">{s["valeur"]}</div><div class="sub">{s["detail"]}</div></div>'
    if "Signal global chantier" in synth:
        s = synth["Signal global chantier"]
        cc = signal_class(s["valeur"])
        synth_html += f'<div class="kpi"><div class="lbl">Signal global</div><div class="val {cc}" style="font-size:18px">{s["valeur"]}</div><div class="sub">{s["detail"]}</div></div>'

    # ── Chart data JSON ───────────────────────────────────────
    chart_labels  = json.dumps([z["zone"].split(" — ")[0] for z in zones])
    chart_cfo     = json.dumps([pct_num(z["cfo"]) for z in zones])
    chart_cfa     = json.dumps([pct_num(z["cfa"]) for z in zones])
    chart_h_theo  = json.dumps([float(str(z["h_cfo"]).replace("h",""))+float(str(z["h_cfa"]).replace("h","")) for z in zones])
    chart_h_point = json.dumps([pct_num(z["h_point"]) for z in zones])
    chart_ecarts  = json.dumps([pct_num(z["ecart"]) for z in zones])

    # ── KPI header ────────────────────────────────────────────
    h_theo_kpi  = kpi.get("h_theo",  f"{data.get('h_totales_terrain',0)}h")
    h_point_kpi = kpi.get("h_point", f"{data.get('h_pointees',0)}h")
    derives_kpi = kpi.get("derives", "—")
    bl_kpi      = kpi.get("bl",      "—")

    # ── H_point color ─────────────────────────────────────────
    try:
        ht = float(str(h_theo_kpi).replace("h",""))
        hp = float(str(h_point_kpi).replace("h",""))
        ep_global = abs((hp-ht)/ht*100) if ht else 0
        h_point_cls = "err" if ep_global > 15 else ("warn" if ep_global > 5 else "ok")
    except:
        h_point_cls = "warn"

    derives_cls = signal_class(derives_kpi)
    bl_cls      = signal_class(bl_kpi)

    html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>EarlySign · Dashboard DR — {data['chantier']}</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600;700;800&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
:root {{
  --p:#5E17EB; --p-d:#311686; --p-m:#4F26DB; --p-l:#EEE6FD;
  --p-10:rgba(94,23,235,.10); --p-06:rgba(94,23,235,.06);
  --y:#FEBE32; --y-d:#C99200; --y-10:rgba(254,190,50,.12);
  --ink:#2A2055; --ink-70:#504880; --ink-40:#9490B0;
  --bg:#F0F2F8; --s1:#F4F5F8; --s2:#ECEEF4; --card:#FFFFFF; --bd:#E4E6EF;
  --ok:#16A34A; --ok-bg:rgba(34,197,94,.08);
  --err:#DC2626; --err-bg:rgba(220,38,38,.08);
  --warn:#D97706; --warn-bg:rgba(217,119,6,.08);
  --font:'Inter',system-ui,sans-serif; --font-head:'Montserrat',sans-serif;
  --r:8px; --r3:4px; --r-pill:999px;
}}
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:var(--font);background:var(--bg);color:var(--ink);min-height:100vh;font-size:13px}}
.nav{{background:var(--card);border-bottom:1px solid var(--bd);position:sticky;top:0;z-index:100}}
.nav-inner{{max-width:1400px;margin:0 auto;padding:0 32px;height:56px;display:flex;align-items:center;justify-content:space-between}}
.hdr-left{{display:flex;align-items:center;gap:14px}}
.logo-seve{{font-weight:800;font-family:var(--font-head);color:var(--ink);font-size:15px;letter-spacing:-.5px}}
.logo-up{{font-weight:800;font-family:var(--font-head);color:var(--y);font-size:9px;vertical-align:super}}
.hdr-sep{{width:1px;height:22px;background:var(--bd)}}
.hdr-cat{{font-size:9px;font-weight:700;letter-spacing:.16em;text-transform:uppercase;color:var(--p)}}
.hdr-title{{font-size:14px;font-weight:700;color:var(--ink);font-family:var(--font-head)}}
.hdr-title span{{color:var(--p)}}
.hdr-right{{display:flex;align-items:center;gap:10px}}
.tag{{background:var(--s1);border:1px solid var(--bd);border-radius:var(--r-pill);padding:4px 12px;font-size:10px;font-weight:600;color:var(--ink-70);display:flex;align-items:center;gap:5px}}
.live-dot{{width:6px;height:6px;border-radius:50%;background:var(--ok);animation:pulse 2.5s infinite}}
@keyframes pulse{{0%,100%{{opacity:1;transform:scale(1)}}50%{{opacity:.5;transform:scale(.8)}}}}
.week-badge{{background:var(--p);color:#fff;border-radius:var(--r-pill);padding:4px 12px;font-size:10px;font-weight:700}}
.main{{max-width:1400px;margin:0 auto;padding:28px 32px}}
.sec{{font-size:9px;font-weight:700;letter-spacing:.18em;text-transform:uppercase;color:var(--ink-40);margin-bottom:12px;display:flex;align-items:center;gap:12px}}
.sec::after{{content:'';flex:1;height:1px;background:var(--bd)}}
.ccard{{background:var(--card);border:1px solid var(--bd);border-radius:var(--r);padding:18px 20px}}
.ccard-t{{font-size:9px;font-weight:700;letter-spacing:.14em;text-transform:uppercase;color:var(--ink-40);margin-bottom:14px}}
.g-hero{{display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:10px;margin-bottom:20px}}
.kpi{{background:var(--card);border:1px solid var(--bd);border-radius:var(--r);padding:16px 18px;transition:border-color .15s;cursor:default}}
.kpi:hover{{border-color:var(--p)}}
.kpi .lbl{{font-size:9px;font-weight:700;letter-spacing:.16em;text-transform:uppercase;color:var(--ink-40);margin-bottom:8px}}
.kpi .val{{font-size:26px;font-weight:600;letter-spacing:-1px;line-height:1;font-family:var(--font-head);color:var(--ink)}}
.kpi .val.accent{{color:var(--p)}} .kpi .val.ok{{color:var(--ok)}} .kpi .val.err{{color:var(--err)}} .kpi .val.warn{{color:var(--warn)}}
.kpi .sub{{font-size:10px;color:var(--ink-40);margin-top:5px}}
.prog-hero{{background:var(--s1);border:1px solid var(--bd);border-radius:var(--r);padding:22px 26px;display:flex;align-items:center;gap:28px;margin-bottom:20px}}
.prog-pct{{font-size:52px;font-weight:600;letter-spacing:-2px;line-height:1;color:var(--p);font-family:var(--font-head);min-width:120px}}
.prog-hero-right{{flex:1}}
.prog-bar-label{{font-size:9px;font-weight:700;letter-spacing:.16em;text-transform:uppercase;color:var(--ink-40);margin-bottom:10px}}
.prog-bar-wrap{{background:var(--s2);border-radius:100px;height:6px;overflow:hidden;margin-bottom:10px}}
.prog-bar{{height:100%;border-radius:100px;background:linear-gradient(90deg,var(--p-m),var(--p));transition:width 1.2s cubic-bezier(.22,1,.36,1);width:0%}}
.prog-sub{{font-size:11px;color:var(--ink-70)}}
.prog-sub b{{color:var(--ink)}}
.zones-table-wrap{{overflow-x:auto;margin-bottom:20px}}
.zones-table{{width:100%;border-collapse:collapse;font-size:11.5px}}
.zones-table th{{background:var(--s1);font-size:9px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;color:var(--ink-40);padding:9px 14px;text-align:left;border-bottom:1px solid var(--bd);white-space:nowrap}}
.zones-table td{{padding:10px 14px;border-bottom:1px solid var(--s2);color:var(--ink-70);vertical-align:middle}}
.zones-table tr:last-child td{{border-bottom:none}}
.zones-table tr:hover td{{background:rgba(94,23,235,.02)}}
.zones-table td:first-child{{font-weight:600;color:var(--ink)}}
.pct-cell{{font-weight:700;font-family:var(--font-head);color:var(--ink)}}
.mini-bar-wrap{{background:var(--s2);border-radius:100px;height:4px;width:80px;overflow:hidden;display:inline-block;vertical-align:middle;margin-left:8px}}
.mini-bar{{height:100%;border-radius:100px;background:var(--p)}}
.sig{{display:inline-flex;align-items:center;gap:4px;font-size:9px;font-weight:700;border-radius:var(--r3);padding:3px 8px;letter-spacing:.06em;text-transform:uppercase}}
.sig.ok{{background:var(--ok-bg);color:var(--ok)}} .sig.err{{background:var(--err-bg);color:var(--err)}} .sig.warn{{background:var(--warn-bg);color:var(--warn)}}
.alertes{{display:flex;flex-direction:column;gap:10px;margin-bottom:20px}}
.alerte{{border-radius:var(--r);padding:14px 16px;display:flex;align-items:flex-start;gap:12px}}
.alerte.err{{background:var(--err-bg);border-left:3px solid var(--err)}}
.alerte.warn{{background:var(--warn-bg);border-left:3px solid var(--warn)}}
.alerte.ok{{background:var(--ok-bg);border-left:3px solid var(--ok)}}
.alerte-icon{{font-size:16px;line-height:1;flex-shrink:0;margin-top:1px}}
.alerte-title{{font-size:11px;font-weight:700;margin-bottom:3px}}
.alerte.err .alerte-title{{color:var(--err)}} .alerte.warn .alerte-title{{color:var(--warn)}} .alerte.ok .alerte-title{{color:var(--ok)}}
.alerte-desc{{font-size:11px;color:var(--ink-70);line-height:1.5}}
.charts2{{display:grid;grid-template-columns:2fr 1fr;gap:12px;margin-bottom:20px}}
.chart-wrap{{position:relative;height:220px}}
.coher-grid{{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:20px}}
.coher-row{{display:flex;align-items:center;justify-content:space-between;padding:9px 0;border-bottom:1px solid var(--s2);font-size:11.5px}}
.coher-row:last-child{{border-bottom:none}}
.coher-label{{color:var(--ink-70);font-weight:500}}
.coher-val{{font-weight:700;font-family:var(--font-head);color:var(--ink);font-size:13px}}
.footer{{text-align:center;padding:24px 0;font-size:10px;color:var(--ink-40)}}
.footer span{{color:var(--p);font-weight:700}}
@media(max-width:900px){{.g-hero{{grid-template-columns:1fr 1fr}}.charts2,.coher-grid{{grid-template-columns:1fr}}}}
@media(max-width:560px){{.main{{padding:16px}}.g-hero{{grid-template-columns:1fr}}.prog-pct{{font-size:40px}}}}
</style>
</head>
<body>
<nav class="nav">
  <div class="nav-inner">
    <div class="hdr-left">
      <div style="display:flex;align-items:baseline;gap:1px">
        <span class="logo-seve">SEVE</span><span class="logo-up">UP</span>
      </div>
      <div class="hdr-sep"></div>
      <div>
        <div class="hdr-cat">EarlySign · Dashboard DR</div>
        <div class="hdr-title"><span>{data['chantier']}</span></div>
      </div>
    </div>
    <div class="hdr-right">
      {f'<span class="week-badge">Semaine {sem}</span>' if sem else ''}
      <span class="tag"><span class="live-dot"></span>Mis à jour le {data['date_saisie']}</span>
    </div>
  </div>
</nav>

<div class="main">

  <div class="sec">Indicateurs clés · {sem_label}</div>
  <div class="g-hero">
    <div class="kpi">
      <div class="lbl">Avancement global</div>
      <div class="val accent" id="kpi-avanc">0%</div>
      <div class="sub">Cumulé depuis démarrage</div>
    </div>
    <div class="kpi">
      <div class="lbl">Heures théoriques</div>
      <div class="val">{h_theo_kpi}</div>
      <div class="sub">Déclaration terrain</div>
    </div>
    <div class="kpi">
      <div class="lbl">Heures pointées</div>
      <div class="val {h_point_cls}">{h_point_kpi}</div>
      <div class="sub">Pointage RH</div>
    </div>
    <div class="kpi">
      <div class="lbl">Zones en dérive</div>
      <div class="val {derives_cls}">{derives_kpi}</div>
      <div class="sub">Alertes actives</div>
    </div>
  </div>

  <div class="prog-hero">
    <div class="prog-pct" id="prog-pct">0%</div>
    <div class="prog-hero-right">
      <div class="prog-bar-label">Avancement global chantier · {sem_label}</div>
      <div class="prog-bar-wrap"><div class="prog-bar" id="prog-bar"></div></div>
      <div class="prog-sub">
        <b>{av}%</b> d'avancement déclaré — {data['chantier']} · <b>{len(zones)}</b> zones suivies
      </div>
    </div>
  </div>

  <div class="sec">Recommandations prioritaires · {sem_label}</div>
  <div class="alertes">{alertes_html}</div>

  <div class="sec">Avancement par zone · {sem_label}</div>
  <div class="ccard" style="margin-bottom:20px">
    <div class="zones-table-wrap">
      <table class="zones-table">
        <thead>
          <tr>
            <th>Zone</th><th>% CFO</th><th>% CFA</th><th>% Moyen</th>
            <th>H. théo CFO</th><th>H. théo CFA</th><th>H. pointées</th>
            <th>Écart</th><th>BL Matériel</th><th>Signal</th>
          </tr>
        </thead>
        <tbody>{zone_rows_html}</tbody>
      </table>
    </div>
  </div>

  <div class="sec">Analyse visuelle · Répartition & cohérence</div>
  <div class="charts2">
    <div class="ccard">
      <div class="ccard-t">Avancement CFO vs CFA par zone</div>
      <div class="chart-wrap"><canvas id="chartZones"></canvas></div>
    </div>
    <div class="ccard">
      <div class="ccard-t">Heures théoriques vs pointées</div>
      <div class="chart-wrap"><canvas id="chartHeures"></canvas></div>
    </div>
  </div>

  <div class="sec">Moteur de cohérence · Croisements objectifs</div>
  <div class="coher-grid">
    <div class="ccard">
      <div class="ccard-t">Contrôle 1 — Heures pointées vs avancement déclaré</div>
      {coher_h_html}
    </div>
    <div class="ccard">
      <div class="ccard-t">Contrôle 2 — Bons de livraison vs quantités posées</div>
      {coher_bl_html}
    </div>
  </div>

  <div class="sec">Synthèse semaine · Signal global chantier</div>
  <div class="g-hero" style="margin-bottom:32px">{synth_html}</div>

  <div class="footer">
    Généré automatiquement par <span>SeveUp · EarlySign</span> · {data['date_saisie']}
  </div>
</div>

<script>
const PROGRESS = {av};
const LABELS   = {chart_labels};
const CFO      = {chart_cfo};
const CFA      = {chart_cfa};
const H_THEO   = {chart_h_theo};
const H_POINT  = {chart_h_point};
const ECARTS   = {chart_ecarts};

setTimeout(() => {{ document.getElementById('prog-bar').style.width = PROGRESS + '%'; }}, 80);
const dur = 1200, t0 = performance.now();
const pctEl = document.getElementById('prog-pct');
const kpiEl = document.getElementById('kpi-avanc');
(function tick(now) {{
  const tt = Math.min((now - t0) / dur, 1);
  const e  = 1 - Math.pow(1 - tt, 3);
  const v  = (PROGRESS * e).toFixed(1) + '%';
  pctEl.textContent = v;
  kpiEl.textContent = v;
  if (tt < 1) requestAnimationFrame(tick);
}})(t0);

new Chart(document.getElementById('chartZones'), {{
  type: 'bar',
  data: {{
    labels: LABELS,
    datasets: [
      {{ label: 'CFO (%)', data: CFO, backgroundColor: 'rgba(94,23,235,.75)', borderRadius: 4 }},
      {{ label: 'CFA (%)', data: CFA, backgroundColor: 'rgba(254,190,50,.75)', borderRadius: 4 }}
    ]
  }},
  options: {{
    responsive: true, maintainAspectRatio: false,
    plugins: {{
      legend: {{ position: 'top', labels: {{ padding: 14, boxWidth: 10, font: {{ size: 10 }} }} }},
      tooltip: {{ backgroundColor: '#2A2055', titleColor: '#fff', bodyColor: '#9490B0', borderColor: '#5E17EB', borderWidth: 1, padding: 10 }}
    }},
    scales: {{
      x: {{ grid: {{ color: 'rgba(228,230,239,.5)' }}, ticks: {{ font: {{ size: 10 }} }} }},
      y: {{ grid: {{ color: 'rgba(228,230,239,.5)' }}, ticks: {{ font: {{ size: 10 }}, callback: v => v + '%' }}, max: 100, min: 0 }}
    }}
  }}
}});

new Chart(document.getElementById('chartHeures'), {{
  type: 'bar',
  data: {{
    labels: LABELS,
    datasets: [
      {{ label: 'Théoriques (h)', data: H_THEO, backgroundColor: 'rgba(94,23,235,.65)', borderRadius: 4 }},
      {{
        label: 'Pointées (h)',
        data: H_POINT,
        backgroundColor: ECARTS.map(e => Math.abs(e) > 15 ? 'rgba(220,38,38,.7)' : Math.abs(e) > 5 ? 'rgba(217,119,6,.7)' : 'rgba(22,163,74,.65)'),
        borderRadius: 4
      }}
    ]
  }},
  options: {{
    responsive: true, maintainAspectRatio: false,
    plugins: {{
      legend: {{ position: 'top', labels: {{ padding: 14, boxWidth: 10, font: {{ size: 10 }} }} }},
      tooltip: {{ backgroundColor: '#2A2055', titleColor: '#fff', bodyColor: '#9490B0', borderColor: '#5E17EB', borderWidth: 1, padding: 10 }}
    }},
    scales: {{
      x: {{ grid: {{ color: 'rgba(228,230,239,.5)' }}, ticks: {{ font: {{ size: 10 }} }} }},
      y: {{ grid: {{ color: 'rgba(228,230,239,.5)' }}, ticks: {{ font: {{ size: 10 }}, callback: v => v + 'h' }} }}
    }}
  }}
}});
</script>
</body>
</html>"""
    return html

# ── GIT PUSH ──────────────────────────────────────────────────
def git_push(root, semaine):
    msg = f"update: Dashboard EarlySign S{semaine} — {today_fr()}" if semaine else f"update: Dashboard EarlySign — {today_fr()}"
    try:
        subprocess.run(["git", "-C", str(root), "add", "index.html"], check=True)
        subprocess.run(["git", "-C", str(root), "commit", "-m", msg], check=True)
        subprocess.run(["git", "-C", str(root), "push"], check=True)
        print(f"🚀  Git push OK — {msg}")
    except subprocess.CalledProcessError as e:
        print(f"⚠️   Git push échoué : {e}")

# ── MAIN ──────────────────────────────────────────────────────
def main():
    xlsx = find_excel()
    data = parse_excel(xlsx)
    html = build_html(data)
    OUTPUT_HTML.write_text(html, encoding="utf-8")
    print(f"✅  Dashboard généré : {OUTPUT_HTML}")
    git_push(OUTPUT_HTML.parent, data.get("semaine",""))
    print("🌐  Netlify déploiement en cours (30s)…")
    print(f"    → https://vocal-mousse-a108a4.netlify.app/")

if __name__ == "__main__":
    main()
